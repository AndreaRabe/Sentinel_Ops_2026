"""Rapports d'activite et exports PDF/Excel.

Les donnees du rapport sont produites une seule fois (`build_report`) puis
rendues dans le format demande : JSON pour l'ecran, XLSX pour l'exploitation,
PDF pour la transmission. Les trois sorties sont donc garanties identiques.

Le filtrage par site suit le scope de l'appelant : un chef d'equipe exporte
ses sites, un responsable l'organisation entiere.
"""

import io
from collections import Counter
from datetime import date, datetime, time, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.enums import TaskStatus
from app.models.user import User
from app.repositories import incident_repository, site_repository, task_repository
from app.services import scope_service

PERIODS = {"day": 1, "week": 7, "month": 30}


def resolve_period(period: str, reference: date) -> tuple[date, date]:
    """Traduit jour/semaine/mois en bornes de dates (cahier des charges section 2)."""
    if period == "day":
        return reference, reference
    if period == "week":
        start = reference - timedelta(days=reference.weekday())
        return start, start + timedelta(days=6)
    if period == "month":
        start = reference.replace(day=1)
        next_month = (start + timedelta(days=32)).replace(day=1)
        return start, next_month - timedelta(days=1)
    raise ValueError(f"Periode inconnue : {period}")


async def build_report(db: AsyncSession, actor: User, start: date, end: date) -> dict:
    site_ids = await scope_service.visible_site_ids(db, actor)
    start_dt = datetime.combine(start, time.min, tzinfo=timezone.utc)
    end_dt = datetime.combine(end, time.min, tzinfo=timezone.utc) + timedelta(days=1)

    tasks = await task_repository.list_for_period(db, start_dt, end_dt, site_ids)
    incidents = await incident_repository.list_for_period(db, start_dt, end_dt, site_ids)

    sites = await site_repository.list_all(db, site_ids=site_ids, include_inactive=True)
    site_names = {site.id: site.name for site in sites}

    status_counts = Counter(str(task.status) for task in tasks)
    completed = status_counts.get(TaskStatus.COMPLETED, 0)
    total = len(tasks)

    return {
        "start": start,
        "end": end,
        "generated_at": datetime.now(timezone.utc),
        "scope": "Tous les sites" if site_ids is None else ", ".join(sorted(site_names.values())),
        "summary": {
            "tasks_total": total,
            "tasks_completed": completed,
            "tasks_late": status_counts.get(TaskStatus.LATE, 0),
            "tasks_cancelled": status_counts.get(TaskStatus.CANCELLED, 0),
            # Taux d'achevement sur la periode, arrondi au dixieme de point.
            "completion_rate": round(completed / total * 100, 1) if total else 0.0,
            "incidents_total": len(incidents),
            "incidents_resolved": sum(
                1 for incident in incidents if incident.resolved_at is not None
            ),
        },
        "tasks_by_status": dict(status_counts),
        "incidents_by_severity": dict(Counter(str(i.severity) for i in incidents)),
        "tasks": [
            {
                "title": task.title,
                "site": site_names.get(task.site_id, "-"),
                "status": str(task.status),
                "priority": str(task.priority),
                "due_at": task.due_at,
                "completed_at": task.completed_at,
            }
            for task in tasks
        ],
        "incidents": [
            {
                "reference": incident.reference,
                "title": incident.title,
                "site": site_names.get(incident.site_id, "-"),
                "severity": str(incident.severity),
                "status": str(incident.status),
                "occurred_at": incident.occurred_at,
                "resolved_at": incident.resolved_at,
            }
            for incident in incidents
        ],
    }


def _format_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def to_excel(report: dict) -> bytes:
    workbook = Workbook()

    summary = workbook.active
    summary.title = "Synthese"
    summary.append(["Rapport Sentinel Ops"])
    summary["A1"].font = Font(bold=True, size=14)
    summary.append(["Periode", f"{report['start']} -> {report['end']}"])
    summary.append(["Perimetre", report["scope"]])
    summary.append(["Genere le", _format_cell(report["generated_at"])])
    summary.append([])
    summary.append(["Indicateur", "Valeur"])
    summary["A6"].font = Font(bold=True)
    summary["B6"].font = Font(bold=True)
    for key, value in report["summary"].items():
        summary.append([key, value])

    for sheet_name, rows, headers in (
        (
            "Taches",
            report["tasks"],
            ["title", "site", "status", "priority", "due_at", "completed_at"],
        ),
        (
            "Incidents",
            report["incidents"],
            ["reference", "title", "site", "severity", "status", "occurred_at", "resolved_at"],
        ),
    ):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in rows:
            sheet.append([_format_cell(row[header]) for header in headers])
        for column, header in enumerate(headers, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = max(
                14, len(header) + 4
            )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _html_table(headers: list[str], rows: list[dict]) -> str:
    head = "".join(f"<th>{header}</th>" for header in headers)
    body = "".join(
        "<tr>" + "".join(f"<td>{_format_cell(row[header])}</td>" for header in headers) + "</tr>"
        for row in rows
    )
    return f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"


def to_pdf(report: dict) -> bytes:
    """Rendu PDF via WeasyPrint a partir d'un HTML autonome.

    Pas de police externe ni d'image distante : le rendu doit rester identique
    sur une machine hors ligne, contrainte du deploiement en LAN.
    """
    from weasyprint import HTML

    summary_rows = "".join(
        f"<tr><td>{key}</td><td class='num'>{value}</td></tr>"
        for key, value in report["summary"].items()
    )
    html = f"""
    <html><head><meta charset="utf-8"><style>
      @page {{ size: A4; margin: 18mm 14mm; }}
      body {{ font-family: sans-serif; font-size: 10pt; color: #16181d; }}
      h1 {{ font-size: 18pt; margin: 0 0 2mm; }}
      .meta {{ color: #5b616e; font-size: 9pt; margin-bottom: 6mm; }}
      h2 {{ font-size: 12pt; margin: 6mm 0 2mm; border-bottom: 1px solid #d6d9e0;
            padding-bottom: 1mm; }}
      table {{ width: 100%; border-collapse: collapse; }}
      th {{ text-align: left; font-size: 8.5pt; text-transform: uppercase;
            letter-spacing: .04em; color: #5b616e; border-bottom: 1px solid #d6d9e0;
            padding: 1.5mm 1mm; }}
      td {{ padding: 1.5mm 1mm; border-bottom: 1px solid #eceef2; }}
      .num {{ font-variant-numeric: tabular-nums; text-align: right; }}
    </style></head><body>
      <h1>Rapport d'activite</h1>
      <div class="meta">
        Periode {report['start']} &rarr; {report['end']} &middot;
        Perimetre : {report['scope']} &middot;
        Genere le {_format_cell(report['generated_at'])}
      </div>
      <h2>Synthese</h2>
      <table><tbody>{summary_rows}</tbody></table>
      <h2>Taches ({len(report['tasks'])})</h2>
      {_html_table(["title", "site", "status", "priority", "due_at", "completed_at"],
                   report["tasks"])}
      <h2>Incidents ({len(report['incidents'])})</h2>
      {_html_table(["reference", "title", "site", "severity", "status", "occurred_at"],
                   report["incidents"])}
    </body></html>
    """
    return HTML(string=html).write_pdf()
